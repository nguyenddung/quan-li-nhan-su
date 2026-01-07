import sqlite3
import os
import shutil
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
UPLOADS_DIR = os.path.join(DATA_DIR, 'uploads')
DB_PATH = os.path.join(DATA_DIR, 'database.sqlite')

os.makedirs(UPLOADS_DIR, exist_ok=True)

def get_conn():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("PRAGMA foreign_keys = ON;")
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS departments (
      id INTEGER PRIMARY KEY,
      name TEXT NOT NULL UNIQUE,
      description TEXT,
      created_at TEXT DEFAULT CURRENT_TIMESTAMP
    );

    CREATE TABLE IF NOT EXISTS members (
      id INTEGER PRIMARY KEY,
      department_id INTEGER NOT NULL,
      full_name TEXT NOT NULL,
      position TEXT,
      email TEXT,
      phone TEXT,
      notes TEXT,
      created_at TEXT DEFAULT CURRENT_TIMESTAMP,
      FOREIGN KEY(department_id) REFERENCES departments(id) ON DELETE CASCADE
    );

    CREATE TABLE IF NOT EXISTS documents (
      id INTEGER PRIMARY KEY,
      member_id INTEGER NOT NULL,
      tt INTEGER,
      so_ky_hieu TEXT,
      ngay_thang TEXT,
      ten_loai_trichyeu TEXT,
      tac_gia TEXT,
      so_to TEXT,
      ghi_chu TEXT,
      file_path TEXT,
      created_at TEXT DEFAULT CURRENT_TIMESTAMP,
      FOREIGN KEY(member_id) REFERENCES members(id) ON DELETE CASCADE
    );
    """)
    conn.commit()
    conn.close()

# ---------------- Departments ----------------
def list_departments():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM departments ORDER BY created_at DESC")
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

def create_department(name, description='', id_=None):
    conn = get_conn()
    cur = conn.cursor()
    if id_:
        cur.execute("INSERT INTO departments (id, name, description) VALUES (?, ?, ?)", (id_, name, description))
        new_id = id_
    else:
        cur.execute("INSERT INTO departments (name, description) VALUES (?, ?)", (name, description))
        new_id = cur.lastrowid
    conn.commit()
    conn.close()
    return new_id

def update_department(id_, name, description):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("UPDATE departments SET name=?, description=? WHERE id=?", (name, description, id_))
    conn.commit()
    conn.close()

def delete_department(id_, remove_files=False):
    """
    Xóa phòng ban và các thành viên + hồ sơ liên quan.
    Nếu remove_files=True -> xóa cả file trên disk (uploads) của những hồ sơ này.
    """
    conn = get_conn()
    cur = conn.cursor()
    # Lấy các member thuộc dept
    cur.execute("SELECT id FROM members WHERE department_id = ?", (id_,))
    member_rows = cur.fetchall()
    member_ids = [r['id'] for r in member_rows]
    if remove_files:
        for mid in member_ids:
            docs = list_documents_by_member(mid)
            for d in docs:
                fp = d.get('file_path')
                if fp and os.path.exists(fp):
                    try:
                        os.remove(fp)
                    except Exception:
                        pass
    # Xóa members (documents bị cascade)
    cur.execute("DELETE FROM members WHERE department_id = ?", (id_,))
    # Xóa department
    cur.execute("DELETE FROM departments WHERE id = ?", (id_,))
    conn.commit()
    conn.close()

# ---------------- Members ----------------
def list_members_by_dept(dept_id, order_by='m.created_at DESC'):
    """
    Lấy members của dept_id. order_by phải là chuỗi SQL hợp lệ cho phần ORDER BY
    (ví dụ: 'm.id ASC', 'm.full_name COLLATE NOCASE ASC', 'm.created_at DESC').
    Trong UI chúng ta chỉ truyền các giá trị định trước.
    """
    conn = get_conn()
    cur = conn.cursor()
    sql = f"SELECT m.*, d.name as department_name FROM members m JOIN departments d ON m.department_id=d.id WHERE department_id=? ORDER BY {order_by}"
    cur.execute(sql, (dept_id,))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

def list_all_members(order_by='m.created_at DESC'):
    conn = get_conn()
    cur = conn.cursor()
    sql = f"SELECT m.*, d.name as department_name FROM members m JOIN departments d ON m.department_id=d.id ORDER BY {order_by}"
    cur.execute(sql)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

def get_member(id_):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM members WHERE id=?", (id_,))
    row = cur.fetchone()
    conn.close()
    return dict(row) if row else None

def create_member(dept_id, full_name, position='', email='', phone='', notes='', id_=None):
    conn = get_conn()
    cur = conn.cursor()
    if id_:
        cur.execute("INSERT INTO members (id, department_id, full_name, position, email, phone, notes) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (id_, dept_id, full_name, position, email, phone, notes))
        new_id = id_
    else:
        cur.execute("INSERT INTO members (department_id, full_name, position, email, phone, notes) VALUES (?, ?, ?, ?, ?, ?)",
                    (dept_id, full_name, position, email, phone, notes))
        new_id = cur.lastrowid
    conn.commit()
    conn.close()
    return new_id

def update_member(id_, full_name, position, email, phone, notes, department_id=None):
    conn = get_conn()
    cur = conn.cursor()
    if department_id:
        cur.execute("""UPDATE members SET full_name=?, position=?, email=?, phone=?, notes=?, department_id=? WHERE id=?""",
                    (full_name, position, email, phone, notes, department_id, id_))
    else:
        cur.execute("""UPDATE members SET full_name=?, position=?, email=?, phone=?, notes=? WHERE id=?""",
                    (full_name, position, email, phone, notes, id_))
    conn.commit()
    conn.close()

def delete_member(id_):
    # delete documents files first
    docs = list_documents_by_member(id_)
    for d in docs:
        if d.get('file_path') and os.path.exists(d['file_path']):
            try:
                os.remove(d['file_path'])
            except Exception:
                pass
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM members WHERE id=?", (id_,))
    conn.commit()
    conn.close()

def change_member_id(old_id, new_id):
    """
    Thay id của member từ old_id -> new_id an toàn (transaction).
    """
    if old_id == new_id:
        return
    conn = get_conn()
    try:
        cur = conn.cursor()
        cur.execute("SELECT * FROM members WHERE id = ?", (old_id,))
        old = cur.fetchone()
        if not old:
            raise ValueError(f"Member id {old_id} không tồn tại")
        cur.execute("SELECT 1 FROM members WHERE id = ?", (new_id,))
        if cur.fetchone():
            raise ValueError(f"Member id {new_id} đã tồn tại")
        conn.execute("BEGIN")
        department_id = old["department_id"]
        full_name = old["full_name"]
        position = old["position"]
        email = old["email"]
        phone = old["phone"]
        notes = old["notes"]
        created_at = old["created_at"]
        cur.execute("""
            INSERT INTO members (id, department_id, full_name, position, email, phone, notes, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (new_id, department_id, full_name, position, email, phone, notes, created_at))
        cur.execute("UPDATE documents SET member_id = ? WHERE member_id = ?", (new_id, old_id))
        cur.execute("DELETE FROM members WHERE id = ?", (old_id,))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

# ---------------- Documents ----------------
def list_documents_by_member(member_id, order_by='tt ASC, created_at DESC'):
    """
    Lấy documents của member_id; order_by là chuỗi SQL hợp lệ cho ORDER BY.
    """
    conn = get_conn()
    cur = conn.cursor()
    sql = f"SELECT * FROM documents WHERE member_id=? ORDER BY {order_by}"
    cur.execute(sql, (member_id,))
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

def get_document(id_):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM documents WHERE id=?", (id_,))
    row = cur.fetchone()
    conn.close()
    return dict(row) if row else None

def _copy_file_to_uploads(src_path):
    if not src_path:
        return None
    if not os.path.exists(src_path):
        return None
    os.makedirs(UPLOADS_DIR, exist_ok=True)
    filename = f"{int(datetime.now().timestamp())}_{os.path.basename(src_path).replace(' ', '_')}"
    dest = os.path.join(UPLOADS_DIR, filename)
    shutil.copy2(src_path, dest)
    return dest

def create_document(member_id, tt=None, so_ky_hieu='', ngay_thang='', ten_loai_trichyeu='',
                    tac_gia='', so_to='', ghi_chu='', file_src_path=None, id_=None):
    file_path = _copy_file_to_uploads(file_src_path) if file_src_path else None
    conn = get_conn()
    cur = conn.cursor()
    if id_:
        cur.execute("""
        INSERT INTO documents (id, member_id, tt, so_ky_hieu, ngay_thang, ten_loai_trichyeu, tac_gia, so_to, ghi_chu, file_path)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (id_, member_id, tt, so_ky_hieu, ngay_thang, ten_loai_trichyeu, tac_gia, so_to, ghi_chu, file_path))
        new_id = id_
    else:
        cur.execute("""
        INSERT INTO documents (member_id, tt, so_ky_hieu, ngay_thang, ten_loai_trichyeu, tac_gia, so_to, ghi_chu, file_path)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (member_id, tt, so_ky_hieu, ngay_thang, ten_loai_trichyeu, tac_gia, so_to, ghi_chu, file_path))
        new_id = cur.lastrowid
    conn.commit()
    conn.close()
    return new_id

def update_document(id_, tt=None, so_ky_hieu='', ngay_thang='', ten_loai_trichyeu='',
                    tac_gia='', so_to='', ghi_chu='', file_src_path=None):
    existing = get_document(id_)
    file_path = existing.get('file_path') if existing else None
    if file_src_path:
        new_path = _copy_file_to_uploads(file_src_path)
        if new_path:
            if file_path and os.path.exists(file_path):
                try:
                    os.remove(file_path)
                except Exception:
                    pass
            file_path = new_path
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
      UPDATE documents SET tt=?, so_ky_hieu=?, ngay_thang=?, ten_loai_trichyeu=?, tac_gia=?, so_to=?, ghi_chu=?, file_path=?
      WHERE id=?
    """, (tt, so_ky_hieu, ngay_thang, ten_loai_trichyeu, tac_gia, so_to, ghi_chu, file_path, id_))
    conn.commit()
    conn.close()

def delete_document(id_):
    doc = get_document(id_)
    if doc and doc.get('file_path') and os.path.exists(doc['file_path']):
        try:
            os.remove(doc['file_path'])
        except Exception:
            pass
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM documents WHERE id=?", (id_,))
    conn.commit()
    conn.close()

# ---------------- Export ZIP ----------------
def export_data_zip(dest_zip_path):
    import zipfile
    with zipfile.ZipFile(dest_zip_path, 'w', zipfile.ZIP_DEFLATED) as z:
        if os.path.exists(DB_PATH):
            z.write(DB_PATH, arcname='database.sqlite')
        for root, dirs, files in os.walk(UPLOADS_DIR):
            for f in files:
                full = os.path.join(root, f)
                arc = os.path.relpath(full, DATA_DIR)
                z.write(full, arcname=os.path.join('uploads', arc))

# ---------------- Search ----------------
def search_members(name_contains=None, position_contains=None, department_id=None, has_docs=None, order_by='m.created_at DESC'):
    conn = get_conn()
    cur = conn.cursor()
    clauses = []
    params = []

    base = "SELECT m.*, d.name as department_name FROM members m JOIN departments d ON m.department_id = d.id"

    if has_docs is True:
        base += " WHERE EXISTS (SELECT 1 FROM documents doc WHERE doc.member_id = m.id)"
    elif has_docs is False:
        base += " WHERE NOT EXISTS (SELECT 1 FROM documents doc WHERE doc.member_id = m.id)"
    else:
        base += " WHERE 1=1"

    if name_contains:
        clauses.append("m.full_name LIKE ?")
        params.append(f"%{name_contains}%")
    if position_contains:
        clauses.append("m.position LIKE ?")
        params.append(f"%{position_contains}%")
    if department_id:
        clauses.append("m.department_id = ?")
        params.append(department_id)

    if clauses:
        base += " AND " + " AND ".join(clauses)

    base += f" ORDER BY {order_by}"

    cur.execute(base, params)
    rows = [dict(r) for r in cur.fetchall()]
    conn.close()
    return rows

# ---------------- Excel Export ----------------
def _safe_sheet_title(name, fallback):
    invalid = set(':\\/ ?*[]')
    s = ''.join(ch for ch in name if ch not in invalid).strip()
    s = s[:28] if len(s) > 28 else s
    if not s:
        s = fallback
    return s

def export_to_excel(dest_xlsx_path):
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        ws_dept = wb['Sheet']
        ws_dept.title = 'Departments'
    else:
        ws_dept = wb.create_sheet('Departments')
    dept_headers = ['id', 'name', 'description', 'created_at']
    ws_dept.append(dept_headers)
    for d in list_departments():
        ws_dept.append([d.get('id'), d.get('name'), d.get('description'), d.get('created_at')])

    ws_mem = wb.create_sheet('Members')
    mem_headers = ['id', 'department_id', 'department_name', 'full_name', 'position', 'email', 'phone', 'notes', 'created_at']
    ws_mem.append(mem_headers)
    members = list_all_members()
    for m in members:
        ws_mem.append([m.get('id'), m.get('department_id'), m.get('department_name'), m.get('full_name'),
                       m.get('position'), m.get('email'), m.get('phone'), m.get('notes'), m.get('created_at')])

    ws_doc = wb.create_sheet('Documents')
    doc_headers = ['id', 'member_id', 'member_name', 'tt', 'so_ky_hieu', 'ngay_thang', 'ten_loai_trichyeu', 'tac_gia', 'so_to', 'ghi_chu', 'file_path', 'created_at']
    ws_doc.append(doc_headers)
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
      SELECT doc.*, m.full_name as member_name
      FROM documents doc
      LEFT JOIN members m ON doc.member_id = m.id
      ORDER BY doc.member_id, doc.tt
    """)
    for r in cur.fetchall():
        rr = dict(r)
        ws_doc.append([rr.get('id'), rr.get('member_id'), rr.get('member_name'), rr.get('tt'),
                       rr.get('so_ky_hieu'), rr.get('ngay_thang'), rr.get('ten_loai_trichyeu'),
                       rr.get('tac_gia'), rr.get('so_to'), rr.get('ghi_chu'), rr.get('file_path'), rr.get('created_at')])
    conn.close()

    # Per-member detail sheets
    for m in members:
        short_name = (m.get('full_name') or '')[:20].replace('/', '_').replace('\\', '_')
        title = _safe_sheet_title(f"{m.get('id')}_{short_name}", f"Member_{m.get('id')}")
        if title in wb.sheetnames:
            title = f"{title}_{m.get('id')}"
        ws = wb.create_sheet(title)
        ws.append(['Field', 'Value'])
        ws.append(['id', m.get('id')])
        ws.append(['full_name', m.get('full_name')])
        ws.append(['position', m.get('position')])
        ws.append(['email', m.get('email')])
        ws.append(['phone', m.get('phone')])
        ws.append(['department', m.get('department_name')])
        ws.append(['notes', m.get('notes')])
        ws.append(['created_at', m.get('created_at')])
        ws.append([])

        docs = list_documents_by_member(m['id'])
        if docs:
            headers = ['tt', 'so_ky_hieu', 'ngay_thang', 'ten_loai_trichyeu', 'tac_gia', 'so_to', 'ghi_chu', 'file_path']
            ws.append(headers)
            for d in docs:
                fp = d.get('file_path') or ''
                abs_fp = os.path.abspath(fp) if fp else ''
                ws.append([d.get('tt'), d.get('so_ky_hieu'), d.get('ngay_thang'), d.get('ten_loai_trichyeu'),
                           d.get('tac_gia'), d.get('so_to'), d.get('ghi_chu'), abs_fp])

    for sheet_name in ['Departments', 'Members', 'Documents']:
        if sheet_name in wb.sheetnames:
            wsx = wb[sheet_name]
            for col in wsx.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        v = str(cell.value) if cell.value is not None else ''
                    except Exception:
                        v = ''
                    if len(v) > max_len:
                        max_len = len(v)
                adjusted = max(10, min(50, max_len + 2))
                wsx.column_dimensions[col_letter].width = adjusted

    wb.save(dest_xlsx_path)

# ---------------- Excel Import ----------------
def _read_sheet_rows(ws):
    rows = []
    headers = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c).strip() if c is not None else '' for c in row]
            continue
        if all(cell is None for cell in row):
            continue
        rec = {}
        for h, cell in zip(headers, row):
            rec[h] = cell
        rows.append(rec)
    return headers, rows

def import_from_excel(src_xlsx_path, clear_first=True):
    if not os.path.exists(src_xlsx_path):
        raise FileNotFoundError("File không tồn tại")
    wb = load_workbook(src_xlsx_path, data_only=True)
    result = {'departments': 0, 'members': 0, 'documents': 0, 'warnings': []}
    if clear_first:
        clear_database(remove_files=False)
    # Departments
    if 'Departments' in wb.sheetnames:
        ws = wb['Departments']
        _, rows = _read_sheet_rows(ws)
        for r in rows:
            try:
                id_ = int(r.get('id')) if r.get('id') not in (None, '') else None
            except Exception:
                id_ = None
            name = r.get('name') or ''
            desc = r.get('description') or ''
            if not name:
                result['warnings'].append(f"Department row missing name: {r}")
                continue
            try:
                create_department(name, desc, id_=id_)
                result['departments'] += 1
            except Exception as e:
                result['warnings'].append(f"Dept insert failed ({name}): {e}")
    # Members
    if 'Members' in wb.sheetnames:
        ws = wb['Members']
        _, rows = _read_sheet_rows(ws)
        for r in rows:
            try:
                id_ = int(r.get('id')) if r.get('id') not in (None, '') else None
            except Exception:
                id_ = None
            dept_id = r.get('department_id')
            try:
                dept_id = int(dept_id) if dept_id not in (None, '') else None
            except Exception:
                dept_id = None
            full_name = r.get('full_name') or ''
            position = r.get('position') or ''
            email = r.get('email') or ''
            phone = r.get('phone') or ''
            notes = r.get('notes') or ''
            if not full_name or not dept_id:
                result['warnings'].append(f"Member missing name or department: {r}")
                continue
            try:
                create_member(dept_id, full_name, position, email, phone, notes, id_=id_)
                result['members'] += 1
            except Exception as e:
                result['warnings'].append(f"Member insert failed ({full_name}): {e}")
    # Documents
    if 'Documents' in wb.sheetnames:
        ws = wb['Documents']
        _, rows = _read_sheet_rows(ws)
        for r in rows:
            try:
                id_ = int(r.get('id')) if r.get('id') not in (None, '') else None
            except Exception:
                id_ = None
            member_id = r.get('member_id')
            try:
                member_id = int(member_id) if member_id not in (None, '') else None
            except Exception:
                member_id = None
            tt = r.get('tt')
            so_ky_hieu = r.get('so_ky_hieu') or ''
            ngay_thang = r.get('ngay_thang') or ''
            ten_loai = r.get('ten_loai_trichyeu') or ''
            tac_gia = r.get('tac_gia') or ''
            so_to = r.get('so_to') or ''
            ghi_chu = r.get('ghi_chu') or ''
            fp = r.get('file_path') or ''
            fp_to_use = None
            if fp:
                possible = str(fp)
                if os.path.exists(possible):
                    try:
                        fp_to_use = _copy_file_to_uploads(possible)
                    except Exception as e:
                        result['warnings'].append(f"Không copy file {possible}: {e}")
                else:
                    alt = os.path.join(DATA_DIR, 'uploads', os.path.basename(possible))
                    if os.path.exists(alt):
                        try:
                            fp_to_use = _copy_file_to_uploads(alt)
                        except Exception as e:
                            result['warnings'].append(f"Không copy file {alt}: {e}")
                    else:
                        result['warnings'].append(f"File hồ sơ không tìm thấy: {possible}")
            if not member_id:
                result['warnings'].append(f"Document row missing member_id: {r}")
                continue
            try:
                create_document(member_id, tt=tt, so_ky_hieu=so_ky_hieu, ngay_thang=ngay_thang,
                                ten_loai_trichyeu=ten_loai, tac_gia=tac_gia, so_to=so_to, ghi_chu=ghi_chu,
                                file_src_path=fp_to_use, id_=id_)
                result['documents'] += 1
            except Exception as e:
                result['warnings'].append(f"Document insert failed (member {member_id}): {e}")
    return result

# ---------------- Clear / Reset DB ----------------
def clear_database(remove_files=False):
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("DELETE FROM documents")
    cur.execute("DELETE FROM members")
    cur.execute("DELETE FROM departments")
    conn.commit()
    conn.close()
    if remove_files:
        for root, dirs, files in os.walk(UPLOADS_DIR):
            for f in files:
                try:
                    os.remove(os.path.join(root, f))
                except Exception:
                    pass

def reset_database(backup_excel_path=None, remove_files=False):
    if backup_excel_path:
        export_to_excel(backup_excel_path)
    clear_database(remove_files=remove_files)